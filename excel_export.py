"""
Excel export module voor VD Vleuten Tijdregistratie
Genereert een gestijld .xlsx bestand met:
  - Tabblad 1: Totaaloverzicht alle medewerkers
  - Tabblad 2+: Eén tabblad per medewerker met dagdetail
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO
import os

# ── VD Vleuten exacte huisstijl kleuren ──────────────────────────────────────
BLUE_DARK   = "1A3A56"   # donkerblauw — logo
BLUE        = "204160"   # blauw — primair
BLUE_LIGHT  = "E8EEF5"   # lichtblauw
BLUE_PALE   = "F3F6F9"   # heel licht blauw
LIME        = "8DC63F"   # limegroen — logo vleugels
LIME_DARK   = "6A9830"   # donker limegroen
LIME_LIGHT  = "F0F7E4"   # licht limegroen
LIME_BORDER = "C8E08A"   # limegroen rand
ORANGE      = "E8650A"   # oranje accent
ORANGE_PALE = "FFF3EC"   # licht oranje
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F5F7F9"
GRAY_BORDER = "D4DDE6"
TEXT_DARK   = "1A2E3D"
TEXT_MID    = "2D4A5E"
TEXT_MUTED  = "5A7A8E"
LATE_RED    = "C0392B"   # te laat rood
LATE_ORANGE = "E8650A"   # te laat oranje
LATE_RED_BG = "FCF0EE"   # licht rood achtergrond
LATE_ORG_BG = "FFF3EC"   # licht oranje achtergrond
# Backwards-compat aliases
GREEN_DARK  = BLUE_DARK
GREEN_MID   = BLUE
GREEN_LIGHT = BLUE_LIGHT
GREEN_PALE  = BLUE_PALE

# ── Border helpers ────────────────────────────────────────────────────────────
def thin(color="C8D8CE"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def top_bottom(color="C8D8CE"):
    s = Side(style='thin', color=color)
    return Border(top=s, bottom=s)

def bottom_only(color="C8D8CE", style='thin'):
    return Border(bottom=Side(style=style, color=color))

def medium_border(color="1A4A2E"):
    s = Side(style='medium', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ── Fill helpers ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── Font helpers ──────────────────────────────────────────────────────────────
def fnt(bold=False, size=10, color=TEXT_DARK, italic=False, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

# ── Alignment helpers ─────────────────────────────────────────────────────────
def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Duration helpers ─────────────────────────────────────────────────────────
def parse_ts(ts):
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(ts, fmt)
        except ValueError:
            continue
    raise ValueError(f"Onbekend tijdformaat: {ts}")

def compute_seconds(entries):
    """Returns (worked_net, break_secs) tuple."""
    worked = breaks = 0
    pending_in = pending_pause = None
    for row in sorted(entries, key=lambda r: r["timestamp"]):
        act = row["action"]
        try:
            ts = parse_ts(row["timestamp"])
        except ValueError:
            continue
        if act == "in":
            pending_in = ts; pending_pause = None
        elif act == "uit" and pending_in:
            worked += (ts - pending_in).total_seconds(); pending_in = None
        elif act == "pauze_in" and pending_in:
            pending_pause = ts
        elif act == "pauze_uit" and pending_pause:
            b = (ts - pending_pause).total_seconds()
            breaks += b; pending_pause = None; pending_in = ts
    net = max(0, worked - breaks)
    return net, breaks

def fmt_dur(seconds):
    if seconds <= 0:
        return "0:00"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    return f"{h}:{m:02d}"

def fmt_dec(seconds):
    return round(seconds / 3600, 2) if seconds > 0 else 0.0

# ── Logo helper ───────────────────────────────────────────────────────────────
def insert_logo(ws, logo_path, anchor="A1"):
    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 200
        img.height = 50
        img.anchor = anchor
        ws.add_image(img)

# ── Column width setter ───────────────────────────────────────────────────────
def set_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ── Section header (dark green banner) ───────────────────────────────────────
def write_header_banner(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = fnt(bold=True, size=11, color=WHITE)
    cell.fill = fill(BLUE_DARK)
    cell.alignment = align('left', 'center')
    cell.border = thin(BLUE_DARK)

# ── Column header row ─────────────────────────────────────────────────────────
def write_col_headers(ws, row, headers, widths=None):
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_i, value=h)
        cell.font = fnt(bold=True, size=9, color=WHITE)
        cell.fill = fill(BLUE)
        cell.alignment = align('center', 'center')
        cell.border = thin(BLUE_DARK)

# ── Main export function ──────────────────────────────────────────────────────
def generate_excel(db_rows, employees_meta, date_from, date_to, logo_path=None):
    # db_rows must include reason column
    """
    db_rows: list of sqlite3.Row with keys: code, name, action, timestamp
    employees_meta: dict code -> name
    date_from, date_to: strings YYYY-MM-DD
    logo_path: path to logo PNG
    Returns BytesIO with the .xlsx file.
    """

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Group data ─────────────────────────────────────────────────────────────
    emp_day = defaultdict(lambda: defaultdict(list))
    for row in db_rows:
        code = row["code"]
        day  = row["timestamp"][:10]
        emp_day[code][day].append(row)

    # ── Compute summary per employee ───────────────────────────────────────────
    summary = []
    for code, name in sorted(employees_meta.items(), key=lambda x: x[1]):
        total_worked = total_breaks = 0
        day_data = []
        for day in sorted(emp_day.get(code, {}).keys()):
            entries = emp_day[code][day]
            worked, brk = compute_seconds(entries)
            total_worked += worked
            total_breaks += brk
            day_data.append({
                "date": day,
                "worked_sec": worked,
                "break_sec": brk,
                "entries": entries,
            })
        summary.append({
            "code": code,
            "name": name,
            "days": day_data,
            "total_worked": total_worked,
            "total_breaks": total_breaks,
        })

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1: Totaaloverzicht
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Totaaloverzicht")
    ws.sheet_view.showGridLines = False
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # Logo area (rows 1-3)
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 14
    insert_logo(ws, logo_path, "A2")

    # Oranje streep onder logo
    ws.merge_cells("A4:J4")
    ws.cell(4, 1).fill = fill(ORANGE)
    ws.row_dimensions[4].height = 4

    # Rapport titel
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 26
    ws.merge_cells("A6:J6")
    c = ws.cell(6, 1, "URENOVERZICHT — TIJDREGISTRATIE")
    c.font = fnt(bold=True, size=14, color=GREEN_DARK)
    c.alignment = align('left', 'center')

    # Periode info
    ws.row_dimensions[7].height = 18
    ws.merge_cells("A7:D7")
    c = ws.cell(7, 1, f"Periode: {date_from}  t/m  {date_to}")
    c.font = fnt(size=10, color=TEXT_MID)
    c.alignment = align('left', 'center')

    ws.merge_cells("F7:J7")
    c = ws.cell(7, 6, f"Gegenereerd op: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
    c.font = fnt(size=9, color=TEXT_MUTED, italic=True)
    c.alignment = align('right', 'center')

    ws.row_dimensions[8].height = 10

    # ── Kolom headers totaaloverzicht ──────────────────────────────────────────
    SUMMARY_HEADERS = [
        "Code", "Naam medewerker", "Gewerkte dagen",
        "Totaal gewerkt", "Totaal pauze", "Netto uren (dec)",
        "Gem. per dag", "Langste dag", ""
    ]
    ws.row_dimensions[9].height = 22
    for ci, h in enumerate(SUMMARY_HEADERS, 1):
        c = ws.cell(9, ci, h)
        c.font = fnt(bold=True, size=9, color=WHITE)
        c.fill = fill(BLUE_DARK)
        c.alignment = align('center', 'center')
        c.border = thin(BLUE_DARK)

    # ── Data rijen totaaloverzicht ─────────────────────────────────────────────
    data_start = 10
    for ri, emp in enumerate(summary):
        row_num = data_start + ri
        ws.row_dimensions[row_num].height = 20

        days_worked = len([d for d in emp["days"] if d["worked_sec"] > 0])
        avg_per_day = emp["total_worked"] / days_worked if days_worked > 0 else 0
        longest_day = max((d["worked_sec"] for d in emp["days"]), default=0)

        row_fill = fill(BLUE_PALE) if ri % 2 == 0 else fill(WHITE)

        values = [
            emp["code"],
            emp["name"],
            days_worked,
            fmt_dur(emp["total_worked"]),
            fmt_dur(emp["total_breaks"]),
            fmt_dec(emp["total_worked"]),
            fmt_dur(avg_per_day),
            fmt_dur(longest_day),
        ]

        for ci, val in enumerate(values, 1):
            c = ws.cell(row_num, ci, val)
            c.fill = row_fill
            c.border = thin(GRAY_BORDER)
            c.alignment = align('center' if ci != 2 else 'left', 'center')
            c.font = fnt(size=10, color=TEXT_DARK, bold=(ci == 2))
            if ci == 6:  # decimale uren
                c.number_format = '0.00'

    # ── Totaalrij ──────────────────────────────────────────────────────────────
    total_row = data_start + len(summary)
    ws.row_dimensions[total_row].height = 22

    total_worked_all = sum(e["total_worked"] for e in summary)
    total_breaks_all = sum(e["total_breaks"] for e in summary)

    ws.merge_cells(f"A{total_row}:B{total_row}")
    c = ws.cell(total_row, 1, "TOTAAL")
    c.font = fnt(bold=True, size=10, color=WHITE)
    c.fill = fill(BLUE)
    c.alignment = align('center', 'center')
    c.border = thin(BLUE_DARK)

    for ci, val in enumerate([
        len(summary),
        sum(len([d for d in e["days"] if d["worked_sec"] > 0]) for e in summary),
        fmt_dur(total_worked_all),
        fmt_dur(total_breaks_all),
        fmt_dec(total_worked_all),
        "", ""
    ], 3):
        c = ws.cell(total_row, ci, val)
        c.font = fnt(bold=True, size=10, color=WHITE)
        c.fill = fill(BLUE)
        c.alignment = align('center', 'center')
        c.border = thin(BLUE_DARK)
        if ci == 8:
            c.number_format = '0.00'

    # ── Legenda ────────────────────────────────────────────────────────────────
    legend_row = total_row + 2
    ws.row_dimensions[legend_row].height = 16
    ws.merge_cells(f"A{legend_row}:J{legend_row}")
    c = ws.cell(legend_row, 1, "Noot: Tijden in formaat U:MM  ·  Netto uren = gewerkt minus pauze  ·  Decimale uren geschikt voor salarisverwerking")
    c.font = fnt(size=9, color=TEXT_MUTED, italic=True)
    c.alignment = align('left', 'center')

    # ── Kolombreedte totaaloverzicht ───────────────────────────────────────────
    set_widths(ws, {
        "A": 10, "B": 26, "C": 16, "D": 16,
        "E": 16, "F": 18, "G": 16, "H": 16, "I": 4
    })

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A10"

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2+: Één tabblad per medewerker
    # ══════════════════════════════════════════════════════════════════════════
    for emp in summary:
        # Gebruik naam als tabbladnaam (max 31 tekens, geen speciale tekens)
        safe_name = emp["name"][:28].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace(":", "")
        ws2 = wb.create_sheet(safe_name)
        ws2.sheet_view.showGridLines = False

        # Logo
        ws2.row_dimensions[1].height = 14
        ws2.row_dimensions[2].height = 30
        ws2.row_dimensions[3].height = 14
        insert_logo(ws2, logo_path, "A2")

        # Oranje streep
        ws2.merge_cells("A4:H4")
        ws2.cell(4, 1).fill = fill(ORANGE)
        ws2.row_dimensions[4].height = 4

        # Titel
        ws2.row_dimensions[5].height = 6
        ws2.row_dimensions[6].height = 26
        ws2.merge_cells("A6:H6")
        c = ws2.cell(6, 1, f"URENKAART — {emp['name'].upper()}")
        c.font = fnt(bold=True, size=13, color=GREEN_DARK)
        c.alignment = align('left', 'center')

        # Medewerker info band
        ws2.row_dimensions[7].height = 20
        ws2.merge_cells("A7:C7")
        c = ws2.cell(7, 1, f"Code: {emp['code']}   ·   Naam: {emp['name']}")
        c.font = fnt(size=10, color=TEXT_MID)
        c.alignment = align('left', 'center')

        ws2.merge_cells("E7:H7")
        c = ws2.cell(7, 5, f"Periode: {date_from} t/m {date_to}   ·   {datetime.now().strftime('%d-%m-%Y %H:%M')}")
        c.font = fnt(size=9, color=TEXT_MUTED, italic=True)
        c.alignment = align('right', 'center')

        ws2.row_dimensions[8].height = 8

        # ── Dagoverzicht header ────────────────────────────────────────────────
        DAY_HEADERS = ["Datum", "Dag", "Inklok", "Uitklok", "Gewerkt", "Pauze", "Netto (dec)", "Reden / Notitie"]
        ws2.row_dimensions[9].height = 22
        for ci, h in enumerate(DAY_HEADERS, 1):
            c = ws2.cell(9, ci, h)
            c.font = fnt(bold=True, size=9, color=WHITE)
            c.fill = fill(BLUE_DARK)
            c.alignment = align('center', 'center')
            c.border = thin(BLUE_DARK)

        # ── Dag rijen ──────────────────────────────────────────────────────────
        DAY_NAMES = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"]
        detail_start = 10

        if not emp["days"]:
            ws2.row_dimensions[detail_start].height = 20
            ws2.merge_cells(f"A{detail_start}:H{detail_start}")
            c = ws2.cell(detail_start, 1, "Geen registraties in deze periode.")
            c.font = fnt(size=10, color=TEXT_MUTED, italic=True)
            c.alignment = align('center', 'center')
            detail_start += 1
        else:
            for di, day in enumerate(emp["days"]):
                row_num = detail_start + di
                ws2.row_dimensions[row_num].height = 19

                dt = datetime.strptime(day["date"], "%Y-%m-%d")
                day_name = DAY_NAMES[dt.weekday()]
                is_weekend = dt.weekday() >= 5

                # Bepaal eerste in en laatste uit
                entries_sorted = sorted(day["entries"], key=lambda e: e["timestamp"])
                first_in  = next((e["timestamp"][11:16] for e in entries_sorted if e["action"] == "in"),  "—")
                first_in_row = next((e for e in entries_sorted if e["action"] == "in"), None)
                first_reason = first_in_row["reason"] if first_in_row and "reason" in first_in_row.keys() and first_in_row["reason"] else None
                last_uit  = next((e["timestamp"][11:16] for e in reversed(entries_sorted) if e["action"] == "uit"), "—")

                # Te laat detectie — reden aanwezig = te laat geweest
                is_late = bool(first_reason)

                # Kleur: te laat = licht oranje, weekend = pale oranje, werkdag alternerend
                if is_late and not is_weekend:
                    row_fill = fill(LATE_ORG_BG)
                elif is_weekend:
                    row_fill = fill(ORANGE_PALE)
                elif di % 2 == 0:
                    row_fill = fill(BLUE_PALE)
                else:
                    row_fill = fill(WHITE)

                row_values = [
                    dt.strftime("%d-%m-%Y"),
                    day_name,
                    first_in,
                    last_uit,
                    fmt_dur(day["worked_sec"]),
                    fmt_dur(day["break_sec"]),
                    fmt_dec(day["worked_sec"]),
                    ""  # notitie kolom
                ]

                for ci, val in enumerate(row_values, 1):
                    c = ws2.cell(row_num, ci, val)
                    c.fill = row_fill
                    c.border = thin(GRAY_BORDER)
                    c.alignment = align('center', 'center')
                    c.font = fnt(size=10, color=TEXT_DARK)
                    if ci == 1:
                        date_color = LATE_RED if is_late else (ORANGE if is_weekend else TEXT_DARK)
                        c.font = fnt(size=10, bold=True, color=date_color)
                    if ci == 2 and is_weekend:
                        c.font = fnt(size=10, color=ORANGE, bold=True)
                    if ci == 7:
                        c.number_format = '0.00'

            detail_start += len(emp["days"])

        # ── Totaalrij per medewerker ───────────────────────────────────────────
        total_row2 = detail_start
        ws2.row_dimensions[total_row2].height = 22

        days_worked = len([d for d in emp["days"] if d["worked_sec"] > 0])

        ws2.merge_cells(f"A{total_row2}:B{total_row2}")
        c = ws2.cell(total_row2, 1, "TOTAAL")
        c.font = fnt(bold=True, size=10, color=WHITE)
        c.fill = fill(BLUE)
        c.alignment = align('center', 'center')
        c.border = thin(BLUE_DARK)

        for ci, val in enumerate(["", "", "", fmt_dur(emp["total_worked"]), fmt_dur(emp["total_breaks"]), fmt_dec(emp["total_worked"]), ""], 3):
            c = ws2.cell(total_row2, ci, val)
            c.font = fnt(bold=True, size=10, color=WHITE)
            c.fill = fill(BLUE)
            c.alignment = align('center', 'center')
            c.border = thin(BLUE_DARK)
            if ci == 8:
                c.number_format = '0.00'

        # ── Samenvattingsblok ─────────────────────────────────────────────────
        summary_start = total_row2 + 2
        ws2.row_dimensions[summary_start].height = 18

        # Groen kader
        summary_items = [
            ("Gewerkte dagen:", str(days_worked)),
            ("Totaal gewerkt:", fmt_dur(emp["total_worked"])),
            ("Totaal pauze:", fmt_dur(emp["total_breaks"])),
            ("Netto uren (dec):", f"{fmt_dec(emp['total_worked']):.2f}"),
            ("Gem. per werkdag:", fmt_dur(emp["total_worked"] / days_worked if days_worked else 0)),
        ]

        for si, (label, val) in enumerate(summary_items):
            r = summary_start + si
            ws2.row_dimensions[r].height = 18
            c = ws2.cell(r, 6, label)
            c.font = fnt(size=10, color=TEXT_MID, bold=True)
            c.alignment = align('right', 'center')
            c.fill = fill(BLUE_PALE)
            c.border = thin(GRAY_BORDER)
            c = ws2.cell(r, 7, val)
            c.font = fnt(size=10, color=GREEN_DARK, bold=True)
            c.alignment = align('center', 'center')
            c.fill = fill(BLUE_PALE)
            c.border = thin(GRAY_BORDER)
            if si == 3:
                c.number_format = '0.00'

        # ── Kolombreedte per medewerker ────────────────────────────────────────
        set_widths(ws2, {
            "A": 14, "B": 6, "C": 10, "D": 10,
            "E": 12, "F": 12, "G": 14, "H": 22
        })
        ws2.freeze_panes = "A10"

    # ── Sla op in BytesIO ──────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
